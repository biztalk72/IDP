"""Export conversation Q&A as Markdown, PDF, Excel, or CSV."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass

from fpdf import FPDF


@dataclass
class QAEntry:
    question: str
    answer: str
    sources: list[dict]  # [{document, page, chunk_text}]


# ---------------------------------------------------------------------------
# Markdown export
# ---------------------------------------------------------------------------

def export_markdown(entries: list[QAEntry], title: str = "IDP Q&A Report") -> str:
    """Generate a Markdown string from a list of Q&A entries."""
    lines: list[str] = [f"# {title}\n"]

    for i, entry in enumerate(entries, 1):
        lines.append(f"## Question {i}")
        lines.append(f"**Q:** {entry.question}\n")
        lines.append(f"### Answer")
        lines.append(entry.answer + "\n")

        if entry.sources:
            lines.append("### Sources")
            for src in entry.sources:
                lines.append(
                    f"- **{src.get('document', 'unknown')}**, "
                    f"Page {src.get('page', '?')}"
                )
            lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

class _IDPPDF(FPDF):
    """Custom FPDF subclass with header/footer."""

    def __init__(self, report_title: str = "IDP Q&A Report"):
        super().__init__()
        self._report_title = report_title

    def header(self):
        self.set_font("Helvetica", "B", 12)
        self.cell(0, 10, self._report_title, align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


def _strip_markdown(text: str) -> str:
    """Lightweight markdown-to-plain-text conversion for PDF cells."""
    text = re.sub(r"#{1,6}\s*", "", text)  # headings
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)  # bold
    text = re.sub(r"\*(.*?)\*", r"\1", text)  # italic
    text = re.sub(r"`(.*?)`", r"\1", text)  # inline code
    text = re.sub(r"\[Source \d+\]", "", text)  # source refs
    return text.strip()


def export_pdf(entries: list[QAEntry], title: str = "IDP Q&A Report") -> bytes:
    """Generate a PDF report and return raw bytes."""
    pdf = _IDPPDF(report_title=title)
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    for i, entry in enumerate(entries, 1):
        # Question
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, f"Q{i}: {entry.question}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Answer
        pdf.set_font("Helvetica", "", 10)
        answer_text = _strip_markdown(entry.answer)
        pdf.multi_cell(0, 6, answer_text)
        pdf.ln(3)

        # Sources
        if entry.sources:
            pdf.set_font("Helvetica", "I", 9)
            for src in entry.sources:
                pdf.cell(
                    0,
                    5,
                    f"  Source: {src.get('document', '?')}, Page {src.get('page', '?')}",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )
            pdf.ln(4)

        # Separator
        pdf.set_draw_color(200, 200, 200)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(6)

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Table extraction / Excel / CSV export
# ---------------------------------------------------------------------------

def extract_markdown_tables(text: str) -> list[list[list[str]]]:
    """Extract markdown tables from text. Returns list of tables, each a list of rows."""
    tables: list[list[list[str]]] = []
    lines = text.split("\n")
    current_table: list[list[str]] = []

    for line in lines:
        stripped = line.strip()
        if "|" in stripped and stripped.startswith("|"):
            # Skip separator rows like |---|---|---|
            if re.match(r"^\|[\s\-:|]+\|$", stripped):
                continue
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            if cells:
                current_table.append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
    if current_table:
        tables.append(current_table)
    return tables


def table_to_csv(table: list[list[str]]) -> str:
    """Convert a table (list of rows) to CSV string."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in table:
        writer.writerow(row)
    return buf.getvalue()


def table_to_excel(table: list[list[str]]) -> bytes:
    """Convert a table (list of rows) to Excel bytes."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for row in table:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
